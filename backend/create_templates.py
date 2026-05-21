"""Generate finding templates: XLSX, CSV, JSON"""
import os, json, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
os.makedirs(TEMPLATE_DIR, exist_ok=True)

# ── Column definitions ────────────────────────────────────────────────────────
HEADERS = [
    "Finding Code",
    "No",
    "Vulnerability",
    "Risk Rating",
    "CVE ID",
    "CVSS Score",
    "Reference",
    "Affected Endpoint",
    "Observation & Implication",
    "Recommendation",
    "Status",
    "Date Found",
    "Due Date",
    "Remark",
]

# Finding Code col is auto-generated — mark as read-only in styling
AUTO_COL = 1   # 1-indexed: Finding Code
REQ_COLS = {3}  # Vulnerability = required

COL_WIDTHS = [18, 5, 38, 14, 16, 10, 22, 36, 52, 52, 14, 12, 12, 26]

# ── Sample data ───────────────────────────────────────────────────────────────
SAMPLE_ROWS = [
    ["PT26_0001_01", 1, "SQL Injection in Login", "High", "CVE-2024-1234", "8.8",
     "CWE-89",
     "POST /api/v1/login\nPOST /api/v1/search",
     "The application does not properly sanitize user-supplied input in the login and search "
     "endpoints. An attacker can inject arbitrary SQL commands to bypass authentication or "
     "extract sensitive data from the database.",
     "1. Use parameterized queries or prepared statements.\n"
     "2. Implement input validation and whitelist allowed characters.\n"
     "3. Apply principle of least privilege to database accounts.",
     "Open", "2026-01-15", "2026-02-15", "Confirmed exploitable in UAT environment"],

    ["PT26_0001_02", 2, "Stored Cross-Site Scripting (XSS)", "High", "", "7.6",
     "CWE-79 / OWASP A03:2021",
     "GET /api/v1/profile?name=\nPOST /api/v1/comment",
     "User-supplied data is reflected in the response without sanitization, allowing persistent "
     "script injection that executes in victim browsers.",
     "1. Encode all user-supplied output using context-aware encoding.\n"
     "2. Implement Content Security Policy (CSP) headers.\n"
     "3. Validate and sanitize all input on the server side.",
     "In Progress", "2026-01-16", "2026-02-20", "Dev team has started patching"],

    ["PT26_0001_03", 3, "Insecure Direct Object Reference (IDOR)", "Medium", "", "6.5",
     "CWE-639 / OWASP A01:2021",
     "GET /api/v1/users/{id}\nGET /api/v1/orders/{id}",
     "The application uses sequential integer IDs without verifying that the requesting user "
     "is authorized to access the resource, allowing horizontal privilege escalation.",
     "1. Implement proper authorization checks on every resource access.\n"
     "2. Use non-guessable identifiers (UUID) instead of sequential IDs.\n"
     "3. Enforce object-level access control (RBAC/ABAC).",
     "Open", "2026-01-17", "2026-03-01", ""],

    ["PT26_0001_04", 4, "Sensitive Data Exposure in API Response", "Low", "", "4.3",
     "CWE-200",
     "GET /api/v1/user/profile",
     "API response includes sensitive internal fields (password_hash, internal_id, api_secret) "
     "that should not be exposed to clients.",
     "1. Implement response filtering to exclude sensitive fields.\n"
     "2. Use dedicated response DTOs/serializers.\n"
     "3. Review all API responses for data over-exposure.",
     "Open", "2026-01-18", "2026-03-15", ""],

    ["PT26_0001_05", 5, "Missing Rate Limiting on Authentication", "Information", "", "",
     "OWASP A05:2021",
     "POST /api/v1/login\nPOST /api/v1/otp/verify",
     "No rate limiting is applied on authentication endpoints, making the application "
     "susceptible to brute-force and credential stuffing attacks.",
     "1. Implement rate limiting (e.g., 5 attempts per minute per IP).\n"
     "2. Add progressive account lockout after N failed attempts.\n"
     "3. Implement CAPTCHA for repeated failures.",
     "Open", "2026-01-18", "", ""],
]

SEV_COLORS = {
    "Critical": "FFEF4444", "High": "FFF97316",
    "Medium": "FFEAB308", "Low": "FF22C55E", "Information": "FFA855F7",
}
STAT_COLORS = {
    "Open": "FFEF4444", "In Progress": "FFF97316", "Closed": "FF22C55E",
}

# ── Excel template ────────────────────────────────────────────────────────────
def create_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Findings"

    thin  = Side(style="thin",   color="FF334155")
    thick = Side(style="medium", color="FF1E293B")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    header_fill  = PatternFill("solid", fgColor="FF1E293B")   # dark slate
    auto_fill    = PatternFill("solid", fgColor="FF1D4ED8")   # blue (auto-gen)
    header_font  = Font(bold=True, color="FFE2E8F0", size=10)
    auto_font    = Font(bold=True, color="FFFFFFFF", size=10)

    # ── Header row ──────────────────────────────────────────────────────────
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col == AUTO_COL:
            cell.fill = auto_fill
            cell.font = auto_font
        elif col in REQ_COLS:
            cell.fill = PatternFill("solid", fgColor="FF0F172A")
            cell.font = Font(bold=True, color="FF60A5FA", size=10)
        else:
            cell.fill = header_fill
            cell.font = header_font

    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    for r, row in enumerate(SAMPLE_ROWS, 2):
        sev  = row[3]   # Risk Rating col index (0-based)
        stat = row[10]  # Status col index (0-based)
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font   = Font(size=10)

            if c == AUTO_COL:
                # Finding Code — blue tint, monospace-like, read-only
                cell.fill = PatternFill("solid", fgColor="FFdbeafe")
                cell.font = Font(size=10, bold=True, color="FF1D4ED8")
                cell.alignment = Alignment(vertical="center", horizontal="center")
            elif c == 4 and sev in SEV_COLORS:
                # Risk Rating — colored by severity
                cell.fill = PatternFill("solid", fgColor=SEV_COLORS[sev])
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 11 and stat in STAT_COLORS:
                # Status — colored
                cell.fill = PatternFill("solid", fgColor=STAT_COLORS[stat])
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 65

    # ── Column widths & freeze ───────────────────────────────────────────────
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"   # freeze Finding Code + No columns

    # ── Instructions sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("VulnTrack — Finding Template Instructions", True, "FF1D4ED8"),
        ("", False, None),
        ("COLUMN GUIDE", True, "FF0F172A"),
        ("Finding Code  [AUTO]  — Generated by VulnTrack (PT26_0001_01). Do not edit.", False, "FF1D4ED8"),
        ("No            — Sequence number. Auto-assigned if blank.", False, None),
        ("Vulnerability — Full vulnerability name. REQUIRED.", False, "FF991B1B"),
        ("Risk Rating   — Critical / High / Medium / Low / Information", False, None),
        ("CVE ID        — Optional CVE identifier (e.g. CVE-2024-1234)", False, None),
        ("CVSS Score    — Numeric score 0.0–10.0", False, None),
        ("Reference     — CWE ID, OWASP category, or custom reference", False, None),
        ("Affected Endpoint — Target URLs, endpoints, or locations (multi-line OK)", False, None),
        ("Observation & Implication — Describe the finding and its business impact", False, None),
        ("Recommendation — Step-by-step remediation guidance", False, None),
        ("Status        — Open / In Progress / Closed", False, None),
        ("Date Found    — YYYY-MM-DD format", False, None),
        ("Due Date      — YYYY-MM-DD format (remediation deadline)", False, None),
        ("Remark        — Additional notes or context", False, None),
        ("", False, None),
        ("IMPORT NOTES", True, "FF0F172A"),
        ("• Column headers are matched case-insensitively — column order does not matter", False, None),
        ("• Finding Code column is ignored on import — system generates its own codes", False, "FF1D4ED8"),
        ("• Findings are matched by Vulnerability name: update if exists, create if new", False, None),
        ("• Rows with blank Vulnerability column are skipped", False, None),
        ("• Risk Rating aliases: critical, high, medium/med, low, information/info", False, None),
        ("• Status aliases: open | in progress | closed / fixed / resolved / done", False, None),
        ("", False, None),
        ("FINDING CODE FORMAT", True, "FF0F172A"),
        ("  {PREFIX}{YY}_{NNNN}_{FF}", False, "FF1D4ED8"),
        ("  PREFIX = project prefix (e.g. PT, VA, RA)  — set per project", False, None),
        ("  YY     = 2-digit year from project Start Date (e.g. 26 = 2026)", False, None),
        ("  NNNN   = 4-digit project sequence, resets each year (0001, 0002 …)", False, None),
        ("  FF     = 2-digit finding sequence within the project (01, 02 …)", False, None),
        ("  Example: PT26_0001_03  =  Pentest 2026, Project #1, Finding #3", False, None),
    ]
    for row_num, (text, bold, color) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=text)
        fc = color if color else "FF0F172A"
        cell.font = Font(bold=bold, size=10 if not bold else 11, color=fc)
        if bold and text:
            cell.fill = PatternFill("solid", fgColor="FFF1F5F9")
    ws2.column_dimensions["A"].width = 75

    path = os.path.join(TEMPLATE_DIR, "finding_template.xlsx")
    try:
        wb.save(path)
        print(f"  Created : {path}")
    except PermissionError:
        print(f"  Skipped (file is open in Excel): {path}")

# ── CSV template ──────────────────────────────────────────────────────────────
def create_csv():
    path = os.path.join(TEMPLATE_DIR, "finding_template.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for row in SAMPLE_ROWS:
            writer.writerow(row)
    print(f"  Created : {path}")

# ── JSON template ─────────────────────────────────────────────────────────────
def create_json():
    records = []
    for row in SAMPLE_ROWS:
        records.append({
            "Finding Code":              row[0],
            "No":                        row[1],
            "Vulnerability":             row[2],
            "Risk Rating":               row[3],
            "CVE ID":                    row[4],
            "CVSS Score":                row[5],
            "Reference":                 row[6],
            "Affected Endpoint":         row[7],
            "Observation & Implication": row[8],
            "Recommendation":            row[9],
            "Status":                    row[10],
            "Date Found":                row[11],
            "Due Date":                  row[12],
            "Remark":                    row[13],
        })
    path = os.path.join(TEMPLATE_DIR, "finding_template.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    print(f"  Created : {path}")


if __name__ == "__main__":
    print("[VulnTrack] Generating templates...")
    create_xlsx()
    create_csv()
    create_json()
    print("[VulnTrack] Done.")
